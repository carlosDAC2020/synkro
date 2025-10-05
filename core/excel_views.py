"""
Vistas para importación y exportación masiva mediante Excel
"""
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Cliente, Producto, Categoria, Proveedor


# ============= CLIENTES =============

@login_required
def clientes_exportar_excel(request):
    """Exporta todos los clientes a un archivo Excel"""
    clientes = Cliente.objects.all().order_by('id')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Teléfono', 'Email', 'Razón Social', 'Dirección']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row, cliente in enumerate(clientes, 2):
        ws.cell(row=row, column=1, value=cliente.id)
        ws.cell(row=row, column=2, value=cliente.nombre)
        ws.cell(row=row, column=3, value=cliente.telefono or '')
        ws.cell(row=row, column=4, value=cliente.email or '')
        ws.cell(row=row, column=5, value=cliente.razon_social or '')
        ws.cell(row=row, column=6, value=cliente.direccion or '')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=clientes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


@login_required
def clientes_descargar_plantilla(request):
    """Descarga una plantilla Excel vacía para importar clientes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Teléfono', 'Email', 'Razón Social', 'Dirección']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Fila de ejemplo
    ejemplo = ['', 'Juan Pérez', '5551234567', 'juan@ejemplo.com', 'Empresa XYZ', 'Calle Principal 123']
    for col, valor in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col, value=valor)
        cell.fill = example_fill
    
    # Instrucciones
    ws.cell(row=4, column=1, value="INSTRUCCIONES:")
    ws.cell(row=5, column=1, value="• Deja el ID vacío para crear nuevos clientes")
    ws.cell(row=6, column=1, value="• Incluye el ID para actualizar clientes existentes")
    ws.cell(row=7, column=1, value="• El Nombre es obligatorio")
    ws.cell(row=8, column=1, value="• Los demás campos son opcionales")
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_clientes.xlsx'
    wb.save(response)
    return response


@login_required
def clientes_importar_excel(request):
    """Importa clientes desde un archivo Excel"""
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        
        try:
            wb = load_workbook(archivo)
            ws = wb.active
            
            creados = 0
            actualizados = 0
            errores = []
            
            # Procesar filas (saltar encabezado)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[1]:  # Si no hay nombre, saltar
                    continue
                
                try:
                    cliente_id = row[0]
                    nombre = row[1]
                    telefono = row[2] or ''
                    email = row[3] or ''
                    razon_social = row[4] or ''
                    direccion = row[5] or ''
                    
                    if cliente_id:
                        # Actualizar cliente existente
                        cliente = Cliente.objects.filter(id=cliente_id).first()
                        if cliente:
                            cliente.nombre = nombre
                            cliente.telefono = telefono
                            cliente.email = email
                            cliente.razon_social = razon_social
                            cliente.direccion = direccion
                            cliente.save()
                            actualizados += 1
                        else:
                            errores.append(f"Fila {row_num}: Cliente con ID {cliente_id} no encontrado")
                    else:
                        # Crear nuevo cliente
                        Cliente.objects.create(
                            nombre=nombre,
                            telefono=telefono,
                            email=email,
                            razon_social=razon_social,
                            direccion=direccion
                        )
                        creados += 1
                        
                except Exception as e:
                    errores.append(f"Fila {row_num}: {str(e)}")
            
            # Mensajes de resultado
            if creados > 0:
                messages.success(request, f'{creados} cliente(s) creado(s) exitosamente.')
            if actualizados > 0:
                messages.success(request, f'{actualizados} cliente(s) actualizado(s) exitosamente.')
            if errores:
                for error in errores[:5]:  # Mostrar solo los primeros 5 errores
                    messages.error(request, error)
                if len(errores) > 5:
                    messages.error(request, f'... y {len(errores) - 5} errores más.')
                    
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return redirect('cliente_list')


# ============= PRODUCTOS =============

@login_required
def productos_exportar_excel(request):
    """Exporta todos los productos a un archivo Excel"""
    productos = Producto.objects.select_related('categoria').order_by('id')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Encabezados
    headers = ['ID', 'Nombre', 'SKU', 'Categoría', 'Descripción', 'Stock Actual', 'Stock Mínimo', 'Costo Unitario', 'Precio Venta']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.id)
        ws.cell(row=row, column=2, value=producto.nombre)
        ws.cell(row=row, column=3, value=producto.sku)
        ws.cell(row=row, column=4, value=producto.categoria.nombre if producto.categoria else '')
        ws.cell(row=row, column=5, value=producto.descripcion or '')
        ws.cell(row=row, column=6, value=producto.stock_actual)
        ws.cell(row=row, column=7, value=producto.stock_minimo)
        ws.cell(row=row, column=8, value=float(producto.costo_unitario))
        ws.cell(row=row, column=9, value=float(producto.precio_venta))
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


@login_required
def productos_descargar_plantilla(request):
    """Descarga una plantilla Excel vacía para importar productos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Encabezados
    headers = ['ID', 'Nombre', 'SKU', 'Categoría', 'Descripción', 'Stock Actual', 'Stock Mínimo', 'Costo Unitario', 'Precio Venta']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Fila de ejemplo
    ejemplo = ['', 'Laptop Dell XPS 13', 'DELL-XPS13', 'Electrónica', 'Laptop ultradelgada', 10, 5, 800.00, 1200.00]
    for col, valor in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col, value=valor)
        cell.fill = example_fill
    
    # Instrucciones
    ws.cell(row=4, column=1, value="INSTRUCCIONES:")
    ws.cell(row=5, column=1, value="• Deja el ID vacío para crear nuevos productos")
    ws.cell(row=6, column=1, value="• Incluye el ID para actualizar productos existentes")
    ws.cell(row=7, column=1, value="• Nombre y SKU son obligatorios")
    ws.cell(row=8, column=1, value="• La Categoría debe existir previamente (nombre exacto)")
    ws.cell(row=9, column=1, value="• Stock, Costo y Precio deben ser números")
    
    # Categorías disponibles
    categorias = Categoria.objects.all()
    if categorias:
        ws.cell(row=11, column=1, value="CATEGORÍAS DISPONIBLES:")
        for idx, cat in enumerate(categorias, 12):
            ws.cell(row=idx, column=1, value=f"• {cat.nombre}")
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_productos.xlsx'
    wb.save(response)
    return response


@login_required
def productos_importar_excel(request):
    """Importa productos desde un archivo Excel"""
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        
        try:
            wb = load_workbook(archivo)
            ws = wb.active
            
            creados = 0
            actualizados = 0
            errores = []
            
            # Procesar filas (saltar encabezado)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[1] or not row[2]:  # Si no hay nombre o SKU, saltar
                    continue
                
                try:
                    producto_id = row[0]
                    nombre = row[1]
                    sku = row[2]
                    categoria_nombre = row[3] or ''
                    descripcion = row[4] or ''
                    stock_actual = int(row[5]) if row[5] else 0
                    stock_minimo = int(row[6]) if row[6] else 0
                    costo_unitario = float(row[7]) if row[7] else 0.0
                    precio_venta = float(row[8]) if row[8] else 0.0
                    
                    # Buscar categoría
                    categoria = None
                    if categoria_nombre:
                        categoria = Categoria.objects.filter(nombre__iexact=categoria_nombre).first()
                    
                    if producto_id:
                        # Actualizar producto existente
                        producto = Producto.objects.filter(id=producto_id).first()
                        if producto:
                            producto.nombre = nombre
                            producto.sku = sku
                            producto.categoria = categoria
                            producto.descripcion = descripcion
                            producto.stock_actual = stock_actual
                            producto.stock_minimo = stock_minimo
                            producto.costo_unitario = costo_unitario
                            producto.precio_venta = precio_venta
                            producto.save()
                            actualizados += 1
                        else:
                            errores.append(f"Fila {row_num}: Producto con ID {producto_id} no encontrado")
                    else:
                        # Crear nuevo producto
                        Producto.objects.create(
                            nombre=nombre,
                            sku=sku,
                            categoria=categoria,
                            descripcion=descripcion,
                            stock_actual=stock_actual,
                            stock_minimo=stock_minimo,
                            costo_unitario=costo_unitario,
                            precio_venta=precio_venta
                        )
                        creados += 1
                        
                except Exception as e:
                    errores.append(f"Fila {row_num}: {str(e)}")
            
            # Mensajes de resultado
            if creados > 0:
                messages.success(request, f'{creados} producto(s) creado(s) exitosamente.')
            if actualizados > 0:
                messages.success(request, f'{actualizados} producto(s) actualizado(s) exitosamente.')
            if errores:
                for error in errores[:5]:
                    messages.error(request, error)
                if len(errores) > 5:
                    messages.error(request, f'... y {len(errores) - 5} errores más.')
                    
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return redirect('producto_list')


# ============= PROVEEDORES =============

@login_required
def proveedores_exportar_excel(request):
    """Exporta todos los proveedores a un archivo Excel"""
    proveedores = Proveedor.objects.all().order_by('id')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Estilos
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Contacto', 'Razón Social', 'Dirección', 'Teléfono', 'Email']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row, proveedor in enumerate(proveedores, 2):
        ws.cell(row=row, column=1, value=proveedor.id)
        ws.cell(row=row, column=2, value=proveedor.nombre)
        ws.cell(row=row, column=3, value=proveedor.contacto or '')
        ws.cell(row=row, column=4, value=proveedor.razon_social or '')
        ws.cell(row=row, column=5, value=proveedor.direccion or '')
        ws.cell(row=row, column=6, value=proveedor.telefono or '')
        ws.cell(row=row, column=7, value=proveedor.email or '')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=proveedores_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


@login_required
def proveedores_descargar_plantilla(request):
    """Descarga una plantilla Excel vacía para importar proveedores"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Estilos
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Contacto', 'Razón Social', 'Dirección', 'Teléfono', 'Email']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Fila de ejemplo
    ejemplo = ['', 'Distribuidora ABC', 'María González', 'ABC Distribuciones S.A.', 'Av. Industrial 456', '5559876543', 'contacto@abc.com']
    for col, valor in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col, value=valor)
        cell.fill = example_fill
    
    # Instrucciones
    ws.cell(row=4, column=1, value="INSTRUCCIONES:")
    ws.cell(row=5, column=1, value="• Deja el ID vacío para crear nuevos proveedores")
    ws.cell(row=6, column=1, value="• Incluye el ID para actualizar proveedores existentes")
    ws.cell(row=7, column=1, value="• El Nombre es obligatorio")
    ws.cell(row=8, column=1, value="• Los demás campos son opcionales")
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_proveedores.xlsx'
    wb.save(response)
    return response


@login_required
def proveedores_importar_excel(request):
    """Importa proveedores desde un archivo Excel"""
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        
        try:
            wb = load_workbook(archivo)
            ws = wb.active
            
            creados = 0
            actualizados = 0
            errores = []
            
            # Procesar filas (saltar encabezado)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[1]:  # Si no hay nombre, saltar
                    continue
                
                try:
                    proveedor_id = row[0]
                    nombre = row[1]
                    contacto = row[2] or ''
                    razon_social = row[3] or ''
                    direccion = row[4] or ''
                    telefono = row[5] or ''
                    email = row[6] or ''
                    
                    if proveedor_id:
                        # Actualizar proveedor existente
                        proveedor = Proveedor.objects.filter(id=proveedor_id).first()
                        if proveedor:
                            proveedor.nombre = nombre
                            proveedor.contacto = contacto
                            proveedor.razon_social = razon_social
                            proveedor.direccion = direccion
                            proveedor.telefono = telefono
                            proveedor.email = email
                            proveedor.save()
                            actualizados += 1
                        else:
                            errores.append(f"Fila {row_num}: Proveedor con ID {proveedor_id} no encontrado")
                    else:
                        # Crear nuevo proveedor
                        Proveedor.objects.create(
                            nombre=nombre,
                            contacto=contacto,
                            razon_social=razon_social,
                            direccion=direccion,
                            telefono=telefono,
                            email=email
                        )
                        creados += 1
                        
                except Exception as e:
                    errores.append(f"Fila {row_num}: {str(e)}")
            
            # Mensajes de resultado
            if creados > 0:
                messages.success(request, f'{creados} proveedor(es) creado(s) exitosamente.')
            if actualizados > 0:
                messages.success(request, f'{actualizados} proveedor(es) actualizado(s) exitosamente.')
            if errores:
                for error in errores[:5]:
                    messages.error(request, error)
                if len(errores) > 5:
                    messages.error(request, f'... y {len(errores) - 5} errores más.')
                    
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return redirect('proveedor_list')
